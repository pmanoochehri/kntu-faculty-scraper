from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.microsoft import EdgeChromiumDriverManager
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import time
import os

# Get the script's directory
script_dir = os.path.dirname(os.path.abspath(__file__))
excel_path = os.path.join(script_dir, "proffesors_info.xlsx")

# Setup Edge driver
driver = webdriver.Edge(service=Service(EdgeChromiumDriverManager().install()))
driver.get("https://ee.kntu.ac.ir/faculty-members")
time.sleep(2)

# Scrape data
data, links = [], []
i, n = 1, 0

while True:
    try:
        element = driver.find_elements(
            By.CSS_SELECTOR, 
            f"#portlet_ir_sain_university_people_UniversityFacultyListPortlet > div > div > div > main > div.row.users > div:nth-child({i}) > div > a"
        )[0]
        
        links.append(element.get_attribute("href"))
        data.append(element.text)
        i += 1
        
    except:
        n += 1
        if n != 2:
            i = 1
            driver.find_elements(
                By.CSS_SELECTOR,
                "#portlet_ir_sain_university_people_UniversityFacultyListPortlet > div > div > div > main > div.text-center > ul > li:nth-child(4)"
            )[0].click()
            time.sleep(1)
        else:
            break

driver.close()

# Process data with pandas
faculty_list = []
cons_email = "@kntu.ac.ir"

for item, link in zip(data, links):
    parts = item.split("\n")
    if len(parts) >= 3:
        degree = parts[0]
        name = parts[1]
        field = parts[2]
    else:
        degree = ""
        name = item
        field = ""
    
    email = link.split("~")[-1] + cons_email if "~" in link else ""
    
    faculty_list.append({
        'Name': name,
        'Degree': degree,
        'Field': field,
        'Email': email,
        'Link': link
    })

# Create DataFrame
df = pd.DataFrame(faculty_list)

# Sort by Field column (A to Z)
df = df.sort_values(by='Field', ascending=True)

# Save to Excel
df.to_excel(excel_path, index=False)

# Open with openpyxl and format
wb = load_workbook(excel_path)
ws = wb.active

# Set sheet view from left to right (horizontal)
ws.sheet_view.rightToLeft = False  # Default is left-to-right

# Add auto filter to all columns
ws.auto_filter.ref = ws.dimensions

# Auto-fit column widths based on content
for column in ws.columns:
    max_length = 0
    column_letter = get_column_letter(column[0].column)
    
    for cell in column:
        try:
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        except:
            pass
    
    adjusted_width = min(max_length + 2, 50)
    ws.column_dimensions[column_letter].width = adjusted_width

# Set alignment for all cells to left-to-right
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal='left', vertical='center')

# Save the file
wb.save(excel_path)

print(f"✅ File saved to: {excel_path}")
print(f"📊 Total faculty members: {len(df)}")
print(f"🔍 Filters applied to all columns")
print(f"📋 Data sorted by Field column (A to Z)")
print(f"📏 Column widths auto-adjusted")
print(f"📐 Sheet set to left-to-right orientation")
print("\nSample data (sorted by Field):")
print(df[['Name', 'Field']].head(10))